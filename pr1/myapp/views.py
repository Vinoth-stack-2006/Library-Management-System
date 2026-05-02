from django.shortcuts import render, redirect
from .models import Book

from django.contrib import messages
from .models import AdminUser
from django.contrib.auth.hashers import make_password, check_password
from django.shortcuts import  get_object_or_404
from .models import Book
from django.contrib.auth import authenticate, login
from django.contrib.auth.hashers import check_password
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import Student,StaffUser 

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import check_password
from .models import Student, StaffUser

from django.contrib import messages
from django.contrib.auth.hashers import check_password
from .models import Student, StaffUser

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import check_password
from .models import Student, StaffUser, HODUser
import requests


from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from .utils import student_token_generator as default_token_generator
from django.contrib.auth.hashers import make_password
from .models import Student, StaffUser, HODUser

# --- LOGIN VIEW (your existing code) ---
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from .utils import student_token_generator as default_token_generator
from django.contrib.auth.hashers import make_password, check_password
from .models import Student, StaffUser, HODUser

def login_view(request):
    if request.method == "POST":
        user_id = request.POST.get("username")
        password = request.POST.get("password")

        # Admin Login
        if user_id == "admini" and password == "ciet":
            request.session['role'] = 'admin'
            request.session['admin_id'] = 'admini'
            return redirect("admin_dashboard")

        # Student Login
        try:
            student = Student.objects.get(student_id=user_id)
            if password == student.student_id:
                request.session['role'] = 'student'
                request.session['student_id'] = student.student_id
                return redirect("student_broad")
            else:
                messages.error(request, "Invalid password for student.")
                return render(request, "login.html")
        except Student.DoesNotExist:
            pass

        # Staff Login
        try:
            staff = StaffUser.objects.get(staff_id=user_id)
            if staff.password == password or check_password(password, staff.password):
                request.session['role'] = 'staff'
                request.session['staff_id'] = staff.staff_id
                return redirect("staff1_dashboard")
            else:
                messages.error(request, "Invalid password for staff.")
                return render(request, "login.html")
        except StaffUser.DoesNotExist:
            pass

        # HOD Login
        try:
            hod = HODUser.objects.get(hod_id=user_id)
            if hod.password == password or check_password(password, hod.password):
                request.session['role'] = 'hod'
                request.session['hod_id'] = hod.hod_id
                return redirect("staff_dashboard")
            else:
                messages.error(request, "Invalid password for HOD.")
                return render(request, "login.html")
        except HODUser.DoesNotExist:
            pass

        messages.error(request, "Invalid ID or password.")
    return render(request, "login.html")

# --- FORGOT PASSWORD VIEW ---
def forgot_password(request):
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        user_obj = None

        # Search user in all models
        try:
            user_obj = Student.objects.get(student_id=user_id)
        except Student.DoesNotExist:
            try:
                user_obj = StaffUser.objects.get(staff_id=user_id)
            except StaffUser.DoesNotExist:
                try:
                    user_obj = HODUser.objects.get(hod_id=user_id)
                except HODUser.DoesNotExist:
                    if user_id == "admini":
                        messages.success(request, f"Admin reset link: /reset-password/admin/123token/")
                        return render(request, "forgot_password.html")
                    else:
                        messages.error(request, "User ID not found.")
                        return render(request, "forgot_password.html")

        # Generate reset token + link
        uid = urlsafe_base64_encode(force_bytes(user_obj.pk))
        token = default_token_generator.make_token(user_obj)
        reset_link = request.build_absolute_uri(f"/reset-password/{uid}/{token}/")

        # Instead of sending email, display link on screen
        messages.success(request, f"Password reset link (for testing): {reset_link}")
        return render(request, "forgot_password.html")

    return render(request, "forgot_password.html")

# --- RESET PASSWORD VIEW ---
def reset_password(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user_obj = None

        for model in [Student, StaffUser, HODUser]:
            try:
                user_obj = model.objects.get(pk=uid)
                break
            except model.DoesNotExist:
                continue
    except Exception:
        user_obj = None

    if user_obj and default_token_generator.check_token(user_obj, token):
        if request.method == "POST":
            new_password = request.POST.get("new_password")
            confirm_password = request.POST.get("confirm_password")

            if new_password == confirm_password:
                user_obj.password = make_password(new_password)
                user_obj.save()
                messages.success(request, "Password successfully reset! You can login now.")
                return redirect('login')
            else:
                messages.error(request, "Passwords do not match.")
        return render(request, "reset_password.html")
    else:
        messages.error(request, "Invalid or expired reset link.")
        return redirect('login')






from django.contrib import messages
from django.contrib.auth.hashers import make_password
from .models import StaffUser

def staff_create(request):
    if request.method == "POST":
        staff_id = request.POST.get("staff_id")
        name = request.POST.get("name")
        email = request.POST.get("email")
        password = request.POST.get("password")

        if StaffUser.objects.filter(staff_id=staff_id).exists():
            messages.error(request, "Staff ID already exists")
        elif StaffUser.objects.filter(email=email).exists():
            messages.error(request, "Email already exists")
        else:
            staff = StaffUser(
                staff_id=staff_id,
                name=name,
                email=email,
                password=make_password(password)  # hash password
            )
            staff.save()
            messages.success(request, "Staff created successfully!")
            return redirect("staff_list")

    return render(request, "staff_create.html")


def staff_list(request):
    staffs = StaffUser.objects.all()
    return render(request, "staff_list.html", {"staffs": staffs})




    
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import check_password
 # import your custom model


def dashboard(request):
 
    return render(request, "dashboard.html")
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from .models import AdminUser

def index(request):
    return render(request,"index.html")


# Page to add a new book
# Page to add a new book
def add_book(request):
    if request.method == "POST":
        scanner = request.POST['scanner']
        title = request.POST['title']
        author = request.POST['author']
        year = int(request.POST['year'])
        publisher = request.POST['publisher']
        abstract = request.POST.get('abstract','') 
        department = request.POST.get('department')
        # If scanner already exists, just update the details
        book, created = Book.objects.get_or_create(
            scanner=scanner,
            defaults={
                'title': title,
                'author': author,
                'year': year,
                'publisher': publisher,
                'abstract': abstract,
                'available': True   # Default new book = Available
            }
        )

        if not created:
            # Already exists → update details only
            book.title = title
            book.author = author
            book.year = year
            book.publisher = publisher
            book.abstract = abstract
            book.department= department
            book.available = True  # Book re-added = available
            book.save()

        return redirect('available')  # redirect to function name

    return render(request, 'add_book.html')

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import  Book   # <-- Import both models




# Show available books
def available_books(request):
    books = Book.objects.all()

    for b in books:
        # ✅ check Borrow table by matching book_id
        b.is_currently_available = not Borrow.objects.filter(
            book_id=b.scanner,
            returned=False    # still not returned
        ).exists()

    return render(request, 'available.html', {'books': books})


# Borrow a book

from django.shortcuts import render, redirect
from django.http import HttpResponse


def scan_page(request):
    return render(request, "scanner.html")

def save_student(request):
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        student_name = request.POST.get("student_name")
        student_email = request.POST.get("student_email")
        student_password = make_password("student_password")  # default password, you can modify
        
        # create StudentUser with only ID and Name
        student = Student(
            student_id=student_id,
            student_name=student_name,
            student_email=student_email,
            student_password=student_password,
            last_login=timezone.now(),
            is_active=True,
        )
        student.save()
       
    return redirect("scan")
def student_detail(request, student_id):
    student = Student.objects.get(id=student_id)
    return render(request, "student_detail.html", {"student": student})

from django.shortcuts import render, redirect

def student_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # check credentials manually
        if username == "admin" and password == "ciet":
            return redirect("scan")   # go to scan page
        else:
            return render(request, "student_login.html", {"error": "Invalid Username or Password"})

    return render(request, "student_login.html")


def books_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # check credentials manually
        if username == "admin" and password == "ciet":
            return redirect("add_book")   # go to scan page
        else:
            return render(request, "book_login.html", {"error": "Invalid Username or Password"})

    return render(request, "book_login.html")

def delete_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # check credentials manually
        if username == "admin" and password == "ciet":
            return redirect("delete_book")   # go to scan page
        else:
            return render(request, "delete_login.html", {"error": "Invalid Username or Password"})
   
    return render(request, "delete_login.html")

def deletes_stud(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # check credentials manually
        if username == "admin" and password == "ciet":
            return redirect("delete_student")   # go to scan page
        else:
            return render(request, "deletes_stud.html", {"error": "Invalid Username or Password"})

    return render(request, "deletes_stud.html")


from datetime import timedelta
from django.utils import timezone
from django.contrib import messages

from .models import Student, StaffUser, HODUser, Borrow

def get_user_info(user_id, user_name):
    """
    Return a dict with:
      - type:    'student' | 'staff' | 'hod'
      - limit:   max books allowed
      - period:  window in days (None for unlimited window)
      - due_days: standard loan period in days
    Or None if no matching user.
    """
    stu = Student.objects.filter(student_id=user_id, student_name=user_name).first()
    if stu:
        return {'type':'student','limit':4,'period':None,'due_days':20}

    stf = StaffUser.objects.filter(staff_id=user_id, name=user_name).first()
    if stf:
        return {'type':'staff','limit':10,'period':180,'due_days':60}

    hod = HODUser.objects.filter(hod_id=user_id, name=user_name).first()
    if hod:
        return {'type':'hod','limit':10,'period':180,'due_days':60}

    return None


from .models import Borrow

# STEP 1: Scan Student
def borrow(request):
    if request.method == "POST":
        user_id   = request.POST.get("student_id", "").strip()
        user_name = request.POST.get("student_name", "").strip()
        info = get_user_info(user_id, user_name)
        if not info:
            return render(request, "borrow.html", {"error": "❌ User not found"})
        # store for scan_book
        request.session["user_info"] = info
        request.session["user_id"]   = user_id
        request.session["user_name"] = user_name
        return redirect("scan_book")
    return render(request, "borrow.html")


def borrow_view(request):
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        student_name = request.POST.get("student_name")

        exists = Student.objects.filter(student_id=student_id, student_name=student_name).exists()

        if exists:
            request.session["student_id"] = student_id
            request.session["student_name"] = student_name
            return redirect("scan_book")
        else:
            return render(request, "borrow.html", {"error": "❌ Student not found in database"})

    return render(request, "borrow_book.html")



from django.utils import timezone
from django.contrib import messages

def scan_book(request):
    info      = request.session.get("user_info")
    user_id   = request.session.get("user_id")
    user_name = request.session.get("user_name")
    if not info or not user_id:
        messages.error(request, "❌ Please scan your ID first!")
        return redirect("borrow")

    if request.method == "POST":
        book_id   = request.POST.get("scanner")
        book_name = request.POST.get("title")

        # 1) If already borrowed by this user, return it
        existing = Borrow.objects.filter(
            student_id=user_id, book_id=book_id, returned=False
        ).first()
        if existing:
            existing.returned    = True
            existing.returned_at = timezone.now()
            existing.borrow_duration_days = (existing.returned_at - existing.borrowed_at).days
            existing.save()
            messages.success(request, f"✅ '{book_name}' returned!")
            return redirect("borrow_list")

        # 2) Count active borrows within window or overall
        if info['period']:
            cutoff = timezone.now() - timedelta(days=info['period'])
            count  = Borrow.objects.filter(
                student_id=user_id, returned=False, borrowed_at__gte=cutoff
            ).count()
        else:
            count = Borrow.objects.filter(student_id=user_id, returned=False).count()

        if count >= info['limit']:
            period = "this semester" if info['period'] else ""
            messages.error(request, f"❌ Limit reached: {info['limit']} books {period}")
            return render(request, "scan_book.html", {"error":"Limit exceeded"})

        # 3) Ensure book isn’t currently borrowed by anyone else
        if Borrow.objects.filter(book_id=book_id, returned=False).exists():
            messages.error(request, f"❌ '{book_name}' is unavailable!")
            return render(request, "scan_book.html", {"error":"Book unavailable"})

        # 4) Create a new borrow record
        new = Borrow.objects.create(
            student_id=user_id,
            student_name=user_name,
            book_id=book_id,
            book_name=book_name,
           
        )
        # Override due_date according to user type
        new.due_date = new.borrowed_at + timedelta(days=info['due_days'])
        new.save()

        messages.success(request, f"✅ '{book_name}' issued to {user_name}!")
        messages.info(request, f"📅 Due: {new.due_date.strftime('%d %b %Y')}")
        return redirect("borrow_list")

    return render(request, "scan_book.html")


def get_student_status(student_id):
    """Get student's borrowing status and fines"""
    active_borrows = Borrow.objects.filter(student_id=student_id, returned=False)
    
    total_fine = 0
    for borrow in active_borrows:
        fine = borrow.calculate_fine()
        total_fine += fine
    
    return {
        'active_count': active_borrows.count(),
        'remaining_slots': 4 - active_borrows.count(),
        'active_borrows': active_borrows,
        'total_fine': total_fine,
        'can_borrow': active_borrows.count() < 4 and total_fine == 0
    }


# STEP 3: Borrow List
def borrow_list(request):
    records = Borrow.objects.all().order_by('-borrowed_at')
    
    # Calculate fines and duration for all records
    for record in records:
        if record.returned and record.returned_at:
            # Calculate actual duration for returned books
            delta = record.returned_at - record.borrowed_at
            record.borrow_duration_days = delta.days
        else:
            # Calculate current duration and fine for active books
            record.calculate_fine()
            current_duration = timezone.now() - record.borrowed_at
            record.current_days = current_duration.days

    return render(request, 'borrow_list.html', {'records': records})


def student_status(request):
    """Student dashboard showing their borrow status"""
    if request.session.get('user_type') != 'student':
        return redirect('login')
    
    student_id = request.session.get('student_id')
    student_name = request.session.get('student_name', 'Student')
    
    if not student_id:
        messages.error(request, "Please login as student first")
        return redirect('login')
    
    # Get student's status
    status = get_student_status(student_id)
    
    context = {
        'student_id': student_id,
        'student_name': student_name,
        **status  # Unpacks all status data
    }
    
    return render(request, 'student_status.html', context)

    

# STEP 4: Mark as Returned
def return_book(request, borrow_id):
    borrow = Borrow.objects.get(id=borrow_id)
    borrow.returned = True   # ✅ FIXED
    borrow.save()
    return redirect("borrow_list")

from django.shortcuts import render
from django.utils import timezone
from .models import Borrow

def search_borrow(request):
    if request.method == "POST":
        entered_id = request.POST.get("student_id").strip()
        records = Borrow.objects.filter(student_id=entered_id).order_by('-borrowed_at')

        if not records.exists():
            return render(request, "search_borrow.html", {
                "error": f"No borrow records found for ID {entered_id}"
            })

        now = timezone.now()
        for r in records:
            if r.returned and r.returned_at:
                # Duration from borrow to return
                r.borrow_duration_days = (r.returned_at - r.borrowed_at).days
            else:
                # Duration from borrow until today
                r.borrow_duration_days = (now - r.borrowed_at).days

        return render(request, "search_borrow.html", {
            "records": records,
            "entered_id": entered_id
        })

    return render(request, "search_borrow.html")


def show(request):
    data=Book.objects.all()
    return render(request,'delete_book.html',{'data':data})

def delete(request,idn):
    obj=Book.objects.get(id=idn)
    obj.delete()
    return redirect('/show/')




def show1(request):
    data=Student.objects.all()
    return render(request,'delete_student.html',{'data':data})

def delete1(request,idn1):
    obj=Student.objects.get(id=idn1)
    obj.delete()
    return redirect('/show1/')




# yourapp/views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from io import BytesIO
import csv, io

from .forms import UploadFileForm
from .models import Book

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from openpyxl import load_workbook
from io import BytesIO
import csv, io

from .forms import UploadFileForm
from .models import Student  # make sure Student is imported

def bulk_upload_students(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            f = request.FILES['file']
            created, updated, skipped = 0, 0, 0
            seen_emails = set()  # Track emails to avoid duplicates within upload batch
            try:
                if f.name.endswith('.xlsx'):
                    wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
                    ws = wb.active
                    rows = ws.iter_rows(values_only=True)
                    headers = [str(h).strip() for h in next(rows)]
                    with transaction.atomic():
                        for row in rows:
                            data = dict(zip(headers, row))
                            sid = str(data.get('student_id')).strip() if data.get('student_id') else None
                            email = str(data.get('student_email')).strip() if data.get('student_email') else None

                            # Skip if ID or email is missing or email is duplicated in this import batch
                            if not sid or not email or email in seen_emails:
                                skipped += 1
                                continue

                            seen_emails.add(email)
                            defaults = {
                                'student_name': data.get('student_name') or '',
                                'student_email': email,
                                # add other fields as needed
                            }

                            obj, created_flag = Student.objects.update_or_create(
                                student_id=sid, defaults=defaults
                            )
                            if created_flag:
                                created += 1
                            else:
                                updated += 1

                elif f.name.endswith('.csv'):
                    text = f.read().decode('utf-8')
                    reader = csv.DictReader(io.StringIO(text))
                    with transaction.atomic():
                        for data in reader:
                            sid = str(data.get('student_id')).strip() if data.get('student_id') else None
                            email = str(data.get('student_email')).strip() if data.get('student_email') else None

                            if not sid or not email or email in seen_emails:
                                skipped += 1
                                continue

                            seen_emails.add(email)
                            defaults = {
                                'student_name': data.get('student_name') or '',
                                'student_email': email,
                                # add other fields as needed
                            }

                            obj, created_flag = Student.objects.update_or_create(
                                student_id=sid, defaults=defaults
                            )
                            if created_flag:
                                created += 1
                            else:
                                updated += 1

                else:
                    messages.error(request, "Upload .xlsx or .csv only")
                    return redirect('bulk_upload_students')

                messages.success(request, f"Students Imported → created:{created}, updated:{updated}, skipped:{skipped}")
                return redirect('dashboard')
            except Exception as e:
                messages.error(request, f"Error: {e}")
                return redirect('bulk_upload_students')
    else:
        form = UploadFileForm()
    return render(request, "bulk_upload_students.html", {"form": form})


# yourapp/views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from io import BytesIO
import csv, io

from .forms import UploadFileForm
from .models import  Book


def bulk_upload_books(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            f = request.FILES['file']
            created, updated, skipped = 0, 0, 0
            try:
                if f.name.endswith('.xlsx'):
                    wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
                    ws = wb.active
                    rows = ws.iter_rows(values_only=True)
                    headers = [str(h).strip().lower() for h in next(rows)]  # normalize headers
                    with transaction.atomic():
                        for row in rows:
                            data = dict(zip(headers, row))
                            scanner = str(data.get('scanner') or '').strip()
                            if not scanner:
                                skipped += 1
                                continue
                            defaults = {
                                'title': data.get('title') or '',
                                'author': data.get('author') or '',
                                'year': data.get('year') or '',
                                'publisher': data.get('publisher') or '',
                                'department': data.get('department') or '',
                                'abstract': data.get('abstract') or '',
                                'available': True,
                            }
                            obj, created_flag = Book.objects.update_or_create(
                                scanner=scanner, defaults=defaults
                            )
                            if created_flag:
                                created += 1
                            else:
                                updated += 1

                elif f.name.endswith('.csv'):
                    text = f.read().decode('utf-8')
                    reader = csv.DictReader(io.StringIO(text))
                    with transaction.atomic():
                        for data in reader:
                            scanner = str(data.get('scanner') or '').strip()
                            if not scanner:
                                skipped += 1
                                continue
                            defaults = {
                                'title': data.get('title') or '',
                                'author': data.get('author') or '',
                                'year': data.get('year') or '',
                                'publisher': data.get('publisher') or '',
                                'department': data.get('department') or '',
                                'abstract': data.get('abstract') or '',
                                'available': True,
                            }
                            obj, created_flag = Book.objects.update_or_create(
                                scanner=scanner, defaults=defaults
                            )
                            if created_flag:
                                created += 1
                            else:
                                updated += 1
                else:
                    messages.error(request, "Upload .xlsx or .csv only")
                    return redirect('bulk_upload_books')

                messages.success(
                    request,
                    f"Books Imported → created:{created}, updated:{updated}, skipped:{skipped}"
                )
                return redirect('dashboard')

            except Exception as e:
                messages.error(request, f"Error: {e}")
                return redirect('bulk_upload_books')
    else:
        form = UploadFileForm()
    return render(request, "bulk_upload_books.html", {"form": form})

from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.http import JsonResponse

def book_abstract(request, pk):
    """
    AJAX view: returns JSON {'html': rendered_html} for the book abstract.
    """
    book = get_object_or_404(Book, pk=pk)
    html = render_to_string('book_abstract.html', {'book': book}, request=request)
    return JsonResponse({'html': html})


from django.shortcuts import render
from django.core.mail import send_mail

def report_issue_view(request):
    if request.method == "POST":
        book_id = request.POST.get("book_id")
        student_name = request.POST.get("student_name")
        issue_type = request.POST.get("issue_type")
        student_email = request.POST.get("student_email")

        # Email subject
        subject = f"📚 Book Issue Reported - ID {book_id}"

        # Plain text fallback
        message = f"""
Dear {student_name},

We received your report.

Book ID: {book_id}
Issue: {issue_type}

Our librarian will review it soon.
"""

        # HTML email
        html_message = f"""
<html>
  <body style="font-family: Arial, sans-serif; line-height:1.6; color: #333;">
    <h2 style="color: #2E86C1;">📚 Book Issue Reported</h2>
    <p>Dear <strong>{student_name}</strong>,</p>
    <p>We received your report. Details are as follows:</p>
    <ul>
      <li><strong>Book ID:</strong> {book_id}</li>
      <li><strong>Issue:</strong> {issue_type}</li>
    </ul>
    <p>Our librarian will review it soon and get back to you.</p>
    <p style="color: #555;">Thank you,<br/>Library Team</p>
  </body>
</html>
"""

        from_email = "sarangvss06@gmail.com"
        recipient_list = ["rahulbalu330@gmail.com"]

        # Send email using Twilio SendGrid
        send_mail(
            subject,
            message,
            from_email,
            recipient_list,
            html_message=html_message
        )

        return render(request, "report_issue.html", {"success": True})

    return render(request, "report_issue.html")


# views.py







from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Admin

# Create Admin
def admin_register(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        if Admin.objects.filter(username=username).exists():
            messages.error(request, "Username already exists")
        else:
            Admin.objects.create(username=username, password=password)
            messages.success(request, "Admin created successfully! Please login.")
            return redirect("super_login")

    return render(request, "admin_register.html")

# Admin Login
def login_vieww(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        try:
            admin = Admin.objects.get(username=username, password=password)
            request.session["admin_id"] = admin.id
            return redirect("dashboard1")
        except Admin.DoesNotExist:
            messages.error(request, "Invalid username or password")

    return render(request, "super_login.html")

def dashboard1 (request):
    return render(request, "dashboard1.html")


from django.shortcuts import render, redirect
from .models import Gate

# Add or update a gate
def add_gate(request):
    if request.method == "POST":
        scanner = request.POST['scanner']
        title = request.POST['title']
        author = request.POST['author']
        year = request.POST['year']
        publisher = request.POST['publisher']
        abstract = request.POST.get('abstract', '') 
        department = request.POST.get('department', '')

        # Create or get existing gate
        gate, created = Gate.objects.get_or_create(
            scanner=scanner,
            defaults={
                'title': title,
                'author': author,
                'year': year,
                'publisher': publisher,
                'abstract': abstract,
                'available': True,
                'department': department
            }
        )

        if not created:
            # Update existing gate
            gate.title = title
            gate.author = author
            gate.year = year
            gate.publisher = publisher
            gate.abstract = abstract
            gate.department = department
            gate.available = True
            gate.save()

        return redirect('avabilable_gate')  # URL name for listing page

    return render(request, 'add_gate.html')


# List all gates
def gate_books(request):
    gates = Gate.objects.all()  # Fetch all records
    return render(request, 'gate_books.html', {'gates': gates})



def gate_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # check credentials manually
        if username == "admin" and password == "ciet":
            return redirect("add_gate")   # go to scan page
        else:
            return render(request, "gate_login.html", {"error": "Invalid Username or Password"})

    return render(request, "gate_login.html")


# Show available books
from django.shortcuts import render
from .models import Gate, Borrow  # make sure Borrow exists

def gate_books(request):
    gates = Gate.objects.all()  # fetch all gates

    for gate in gates:
        # Check if the gate book is currently available
        gate.is_currently_available = not Borrow.objects.filter(
            book_id=gate.scanner,
            returned=False
        ).exists()

    return render(request, 'avabilable_gate.html', {'books': gates})


def borrow_gate(request):
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        student_name = request.POST.get("student_name")

        exists = Student.objects.filter(student_id=student_id, student_name=student_name).exists()

        if exists:
            request.session["student_id"] = student_id
            request.session["student_name"] = student_name
            return redirect("scan_gate")
        else:
            return render(request, "borrow.html", {"error": "❌ Student not found in database"})

    return render(request, "borrow_gate.html")
# myapp/views.py
from .models import IssuedBook   # ✅ add this at the top


def scan_gate(request):
    if request.method == "POST":
        book_id = request.POST.get("scanner")
        book_name = request.POST.get("title")

        student_id = request.session.get("student_id")
        student_name = request.session.get("student_name")

        # 🔹 Check if same student already borrowed the same book
        record = IssuedBook.objects.filter(
            student_id=student_id,
            book_id=book_id,
            returned=False
        ).first()

        if record:
            # ✅ Mark as returned instead of creating new row
            record.returned = True
            record.returned_at = timezone.now()
            record.save()
        else:
            # ✅ Create new borrow record
            IssuedBook.objects.create(
                student_id=student_id,
                student_name=student_name,
                book_id=book_id,
                book_name=book_name
            )

        return redirect("borrow_listt")

    return render(request, "scan_gate.html")
def borrow_listt(request):
    records = IssuedBook.objects.all().order_by('borrowed_at')
    
    # Dynamically compute borrow duration in days
    for b in records:
        if b.returned and b.returned_at:
            delta = b.returned_at - b.borrowed_at
            b.borrow_duration_days = delta.days
        else:
            b.borrow_duration_days = None

    return render(request, 'borrow_listt.html', {'records': records})

def return_book(request, borrow_id):
    IssuedBook = IssuedBook.objects.get(id=borrow_id)
    IssuedBook.returned = True   # ✅ FIXED
    IssuedBook.save()
    return redirect("borrow_listt")


from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import NewspaperForm
from .models import Newspaper

# Custom check for admin login
def check_admin_session(request):
    return request.session.get('admin_logged_in', False)

def newspaper_upload(request):
    if request.method == "POST":
        form = NewspaperForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Newspaper uploaded successfully!")
            return redirect("newspaper_upload")  # stay on upload page
    else:
        form = NewspaperForm()

    newspapers = Newspaper.objects.all().order_by("-publish_date")
    return render(request, "newspaper_upload.html", {"form": form, "newspapers": newspapers})



# List view (all users can see)
def newspaper_list(request):
    newspapers = Newspaper.objects.all().order_by("-publish_date")
    return render(request, "newspaper_list.html", {"newspapers": newspapers})

# Delete newspaper (admin only)
def newspaper_delete(request, pk):
    if not check_admin_session(request):
        return redirect('admin_login')

    newspaper = Newspaper.objects.get(pk=pk)
    newspaper.file.delete()  # remove file from storage
    newspaper.delete()
    messages.success(request, "🗑 Newspaper deleted successfully!")
    return redirect("newspaper_upload")

    
def student_broad(request):

    return render(request, "student_broad.html")
def report_issue(request):
    return render(request, "report_issue.html")

def admin_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # check credentials manually
        if username == "admin" and password == "ciet":
            return redirect("staff_create")   # go to scan page
        else:
            return render(request, "book_login.html", {"error": "Invalid Username or Password"})

    return render(request, "admin_login.html")


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.core.paginator import Paginator
from django.utils.timezone import localtime
from django.http import JsonResponse
from .forms import PurchaseRequestForm
from .models import PurchaseRequest, Book, StaffUser, HODUser, HODMessage

# ----------------------------
# HOD purchase request page (UPDATED - now HOD only)
# ----------------------------
def request_purchase(request):
    if request.method == "POST":
        form = PurchaseRequestForm(request.POST)
        if form.is_valid():
            purchase_request = form.save(commit=False)
            
            # Get HOD info from session (changed from staff_id to hod_id)
            hod_id = request.session.get("hod_id")  # Changed
            hod_user = None
            if hod_id:
                hod_user = HODUser.objects.filter(hod_id=hod_id).first()  # Changed
                purchase_request.hod_user = hod_user  # Changed
                # Set the hod_name field
                if hod_user:
                    purchase_request.hod_name = hod_user.name  # Changed
            
            purchase_request.save()
            messages.success(request, "Your purchase request has been submitted!")
            return redirect("staff_dashboard")  # HOD dashboard
        else:
            messages.error(request, "Please fill in all required fields.")
    else:
        form = PurchaseRequestForm()
    return render(request, "request_purchase.html", {"form": form})

# ----------------------------
# HOD dashboard with notifications (UPDATED)
# ----------------------------
def staff_dashboard(request):
    """This is now HOD dashboard with full features"""
    if request.session.get('role') != 'hod':  # Changed role check
        return redirect('login')
    
    hod_id = request.session.get("hod_id")  # Changed from staff_id
    hod_user = HODUser.objects.filter(hod_id=hod_id).first()  # Changed
    notifications = HODMessage.objects.filter(hod=hod_user, read=False).order_by('-created_at') if hod_user else []  # Changed

    # Convert times to IST
    for n in notifications:
        n.ist_time = localtime(n.created_at)

    return render(request, "staff_dashboard.html", {"notifications": notifications})

# ----------------------------
# Fetch unread notifications for HOD (UPDATED)
# ----------------------------
def get_unread_notifications(request):
    hod_id = request.session.get("hod_id")  # Changed
    if not hod_id:
        return JsonResponse({"notifications": []})
    
    hod = HODUser.objects.filter(hod_id=hod_id).first()  # Changed
    if not hod:
        return JsonResponse({"notifications": []})

    messages_qs = HODMessage.objects.filter(hod=hod, read=False).order_by('-created_at')  # Changed
    data = [
        {
            "id": m.id,
            "message": m.message,
            "created_at": localtime(m.created_at).strftime("%d %b %Y %H:%M")
        }
        for m in messages_qs
    ]
    return JsonResponse({"notifications": data})

# ----------------------------
# Mark HOD notifications as read (UPDATED)
# ----------------------------
def mark_notifications_read(request):
    if request.method == "POST":
        hod_id = request.session.get("hod_id")  # Changed
        hod = HODUser.objects.filter(hod_id=hod_id).first()  # Changed
        if hod:
            HODMessage.objects.filter(hod=hod, read=False).update(read=True)  # Changed
    return redirect("staff_dashboard")

# ----------------------------
# Approve request - send notification to HOD (UPDATED)
# ----------------------------
def approve_request(request, pk):
    req = get_object_or_404(PurchaseRequest, pk=pk)
    req.status = "Approved"
    req.approved_by = request.session.get("admin", "Admin")
    req.approved_date = timezone.now()
    req.save()
    
    hod_user = req.hod_user  # Changed from staff_user
    if hod_user:
        HODMessage.objects.create(  # Changed from StaffMessage
            hod=hod_user,  # Changed from staff
            message=f"Your purchase request for '{req.book_title}' has been approved."
        )
    return redirect("purchase_requests_list")

# ----------------------------
# Reject request - send notification to HOD (UPDATED)
# ----------------------------
def reject_request(request, pk):
    req = get_object_or_404(PurchaseRequest, pk=pk)
    req.status = "Rejected"
    req.approved_by = request.session.get("admin", "Admin")
    req.approved_date = timezone.now()
    req.save()
    
    hod_user = req.hod_user  # Changed from staff_user
    if hod_user:
        HODMessage.objects.create(  # Changed from StaffMessage
            hod=hod_user,  # Changed from staff
            message=f"Your purchase request for '{req.book_title}' has been rejected."
        )
    messages.error(request, f"{req.book_title} has been rejected.")
    return redirect("purchase_requests_list")

# ----------------------------
# Regular Staff Dashboard (UNCHANGED - this is perfect)
# ----------------------------
def staff1_dashboard(request):
    """
    Simple dashboard for regular staff members
    - No notification features
    - No purchase request functionality  
    - Basic access to view books and newspapers only
    """

    
    staff_id = request.session.get("staff_id")
    staff_name = request.session.get("staff_name", "Staff")
    
    context = {
        'staff_id': staff_id,
        'staff_name': staff_name,
    }
    
    return render(request, 'staff1_dashboard.html', context)


# ----------------------------
# Admin/Purchase Request Views (UNCHANGED - these are perfect)
# ----------------------------
def view_purchase_requests(request):
    requests_list = PurchaseRequest.objects.all().order_by('-request_date')

    # Status filter (optional)
    status_filter = request.GET.get('status')
    if status_filter in ['Pending', 'Approved', 'Rejected']:
        requests_list = requests_list.filter(status=status_filter)

    # Pagination
    paginator = Paginator(requests_list, 10)  # 10 requests per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'purchase_requests_list.html', {
        'requests': page_obj,
        'status_filter': status_filter or ''
    })

def purchase_requests_list(request):
    status_filter = request.GET.get('status', '')

    if status_filter:
        requests_qs = PurchaseRequest.objects.filter(status=status_filter).order_by('-request_date')
    else:
        requests_qs = PurchaseRequest.objects.all().order_by('-request_date')

    paginator = Paginator(requests_qs, 10)  # Show 10 requests per page
    page_number = request.GET.get('page')
    requests_page = paginator.get_page(page_number)

    return render(request, 'purchase_requests_list.html', {
        'requests': requests_page,
        'status_filter': status_filter})

# ----------------------------
# Exam Views (UNCHANGED - these are perfect)
# ----------------------------
from django.shortcuts import render
from .models import MainExam


def exam_dashboard(request):
    exams = MainExam.objects.all()  # Fetch all MainExam entries
    return render(request, 'exam_dashboard.html', {'exams': exams})

def subbranches(request, main_exam_id):
    exam = get_object_or_404(MainExam, id=main_exam_id)
    subbranches = SubExam.objects.filter(main_exam=exam)
    return render(request, 'subbranches.html', {'exam': exam, 'subbranches': subbranches})

def subbranch_details(request, sub_exam_id):
    subbranch = get_object_or_404(SubExam, id=sub_exam_id)
    return render(request, 'subbranch_details.html', {'subbranch': subbranch})

def admin_dashboard(request):
 
    return render(request, 'admin_dashboard.html')

# Add these HOD views to your views.py

from .models import HODUser

# ----------------------------
# HOD Create
# ----------------------------
def hod_create(request):
    if request.method == "POST":
        hod_id = request.POST.get("hod_id")
        name = request.POST.get("name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        department = request.POST.get("department", "")

        # Check if HOD ID already exists
        if HODUser.objects.filter(hod_id=hod_id).exists():
            messages.error(request, f"HOD with ID '{hod_id}' already exists!")
            return render(request, "hod_create.html")

        try:
            hod = HODUser.objects.create(
                hod_id=hod_id,
                name=name,
                email=email,
                password=password,  # Will be hashed automatically in model
                department=department
            )
            messages.success(request, f"HOD '{name}' created successfully!")
            return redirect("hod_list")  # Redirect to HOD list
        except Exception as e:
            messages.error(request, f"Error creating HOD: {str(e)}")

    return render(request, "hod_create.html")

# ----------------------------
# HOD List
# ----------------------------
def hod_list(request):
    hods = HODUser.objects.all().order_by('hod_id')
    return render(request, "hod_list.html", {"hods": hods})
from django.shortcuts import render 
import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import base64

def generate_barcode_view(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        barcode_type = request.POST.get('barcode_type')

        context = {
            'barcode_type': barcode_type,
        }

        if barcode_type == 'single':
            number = request.POST.get('barcode_number')
            if number:
                # generate barcode
                CODE128 = barcode.get_barcode_class('code128')
                rv = BytesIO()
                CODE128(number, writer=ImageWriter()).write(rv)
                barcode_image_url = 'data:image/png;base64,' + base64.b64encode(rv.getvalue()).decode('utf-8')

                # save to database if needed
                # Barcode.objects.create(title=title, number=number)

                context.update({
                    'barcode_number': number,
                    'barcode_image_url': barcode_image_url
                })
        else:
            start = request.POST.get('start_number')
            end = request.POST.get('end_number')
            barcodes = []
            if start and end:
                for i in range(int(start), int(end)+1):
                    code = str(i)
                    CODE128 = barcode.get_barcode_class('code128')
                    rv = BytesIO()
                    CODE128(code, writer=ImageWriter()).write(rv)
                    image_url = 'data:image/png;base64,' + base64.b64encode(rv.getvalue()).decode('utf-8')
                    barcodes.append((code, image_url))
                    # save to database if needed
                    # Barcode.objects.create(title=title, number=code)
                context['barcodes'] = barcodes
                context['start_number'] = start
                context['end_number'] = end

        return render(request, 'barcode_result.html', context)
    else:
        return render(request, 'generate_barcode.html')